"""Импорт справочников и начальных остатков из файла.

Импорт двухшаговый: сначала предпросмотр с разбором ошибок, потом применение.
Загружать сотни строк вслепую и потом выяснять, что получилось, — верный способ
испортить справочник.

Совпадение ищется по коду: повторная загрузка того же файла обновляет записи,
а не плодит дубли.
"""

import csv
import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import (
    Counterparty,
    Nomenclature,
    Outlet,
    OutletType,
    ProductCategory,
    UnitOfMeasure,
    Warehouse,
)


class ImportKind(StrEnum):
    NOMENCLATURE = "nomenclature"
    COUNTERPARTIES = "counterparties"
    OUTLETS = "outlets"
    STOCK = "stock"


KIND_TITLES: dict[ImportKind, str] = {
    ImportKind.NOMENCLATURE: "Номенклатура",
    ImportKind.COUNTERPARTIES: "Контрагенты",
    ImportKind.OUTLETS: "Торговые точки",
    ImportKind.STOCK: "Начальные остатки",
}


@dataclass(slots=True)
class Column:
    key: str
    title: str
    required: bool = False
    hint: str | None = None


COLUMNS: dict[ImportKind, list[Column]] = {
    ImportKind.NOMENCLATURE: [
        Column("code", "Код", hint="По нему ищется совпадение при повторной загрузке"),
        Column("name", "Наименование", required=True),
        Column("article", "Артикул"),
        Column("category", "Группа", hint="Создаётся, если такой ещё нет"),
        Column("unit", "Единица", hint="кг, шт, упак"),
        Column("price", "Цена"),
        Column("vat_rate", "НДС, %"),
        Column("is_weight_goods", "Весовой", hint="да / нет"),
        Column("barcode", "Штрихкод"),
    ],
    ImportKind.COUNTERPARTIES: [
        Column("code", "Код"),
        Column("name", "Наименование", required=True),
        Column("bin_iin", "БИН/ИИН"),
        Column("address", "Адрес"),
        Column("phone", "Телефон"),
        Column("email", "Почта"),
        Column("contact_person", "Контактное лицо"),
    ],
    ImportKind.OUTLETS: [
        Column("code", "Код"),
        Column("name", "Наименование", required=True),
        Column("counterparty", "Контрагент", required=True, hint="Код или наименование"),
        Column("outlet_type", "Тип точки", hint="Магазин, Супермаркет, Рынок…"),
        Column("address", "Адрес"),
        Column("phone", "Телефон"),
        Column("latitude", "Широта", hint="Для проверки визитов на месте"),
        Column("longitude", "Долгота"),
    ],
    ImportKind.STOCK: [
        Column("warehouse", "Склад", required=True, hint="Код или наименование"),
        Column("nomenclature", "Номенклатура", required=True, hint="Код или наименование"),
        Column("quantity", "Количество", required=True),
        Column("price", "Цена"),
    ],
}


@dataclass(slots=True)
class RowResult:
    """Разобранная строка файла с ошибками, если они есть."""

    line: int
    values: dict[str, Any]
    errors: list[str] = field(default_factory=list)
    action: str = "create"

    @property
    def valid(self) -> bool:
        return not self.errors


def read_table(content: bytes, filename: str) -> list[dict[str, str]]:
    """Читает xlsx или csv в список словарей по заголовкам первой строки."""

    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _read_excel(content)
    return _read_csv(content)


def _read_excel(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        header = [str(cell or "").strip() for cell in next(rows)]
    except StopIteration:
        return []

    result: list[dict[str, str]] = []
    for row in rows:
        values = {
            header[index]: ("" if cell is None else str(cell).strip())
            for index, cell in enumerate(row)
            if index < len(header) and header[index]
        }
        if any(values.values()):
            result.append(values)
    return result


def _read_csv(content: bytes) -> list[dict[str, str]]:
    # Excel сохраняет CSV с BOM и точкой с запятой — читаем именно так.
    text = content.decode("utf-8-sig", errors="replace")
    delimiter = ";" if text.count(";") >= text.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [
        {(key or "").strip(): (value or "").strip() for key, value in row.items()}
        for row in reader
        if any((value or "").strip() for value in row.values())
    ]


def _decimal(value: str, field_name: str, errors: list[str]) -> Decimal | None:
    if not value:
        return None
    try:
        # В выгрузках из Excel дробная часть часто через запятую.
        return Decimal(value.replace(" ", "").replace("\xa0", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        errors.append(f"{field_name}: «{value}» — не число")
        return None


def _boolean(value: str) -> bool | None:
    if not value:
        return None
    return value.strip().lower() in {"да", "yes", "true", "1", "+", "истина"}


def _by_code_or_name(db: Session, model: Any, value: str) -> Any:
    """Ищет элемент справочника сначала по коду, затем по наименованию."""

    if not value:
        return None
    found = db.execute(select(model).where(model.code == value)).unique().scalar_one_or_none()
    if found is not None:
        return found
    return (
        db.execute(select(model).where(func.lower(model.name) == value.lower()))
        .unique()
        .scalars()
        .first()
    )


def parse(db: Session, kind: ImportKind, raw_rows: list[dict[str, str]]) -> list[RowResult]:
    """Разбирает и проверяет строки, ничего не записывая в базу."""

    columns = COLUMNS[kind]
    titles = {column.title.lower(): column.key for column in columns}
    results: list[RowResult] = []

    for index, raw in enumerate(raw_rows, start=2):
        values: dict[str, Any] = {}
        errors: list[str] = []

        # Заголовки сопоставляем без учёта регистра и лишних пробелов.
        normalized = {
            (key or "").strip().lower(): (value or "").strip() for key, value in raw.items()
        }
        for title, key in titles.items():
            values[key] = normalized.get(title, "")

        for column in columns:
            if column.required and not values.get(column.key):
                errors.append(f"{column.title}: обязательное поле")

        action = "create"

        if kind is ImportKind.NOMENCLATURE:
            values["price"] = _decimal(values.get("price", ""), "Цена", errors)
            values["vat_rate"] = _decimal(values.get("vat_rate", ""), "НДС", errors)
            values["is_weight_goods"] = _boolean(values.get("is_weight_goods", ""))
            if values.get("code"):
                exists = _by_code_or_name(db, Nomenclature, values["code"])
                action = "update" if exists else "create"

        elif kind is ImportKind.COUNTERPARTIES:
            if values.get("code"):
                exists = _by_code_or_name(db, Counterparty, values["code"])
                action = "update" if exists else "create"

        elif kind is ImportKind.OUTLETS:
            counterparty = _by_code_or_name(db, Counterparty, values.get("counterparty", ""))
            if values.get("counterparty") and counterparty is None:
                errors.append(f"Контрагент «{values['counterparty']}» не найден")
            values["latitude"] = _decimal(values.get("latitude", ""), "Широта", errors)
            values["longitude"] = _decimal(values.get("longitude", ""), "Долгота", errors)
            if values.get("code"):
                exists = _by_code_or_name(db, Outlet, values["code"])
                action = "update" if exists else "create"

        elif kind is ImportKind.STOCK:
            warehouse = _by_code_or_name(db, Warehouse, values.get("warehouse", ""))
            if values.get("warehouse") and warehouse is None:
                errors.append(f"Склад «{values['warehouse']}» не найден")
            product = _by_code_or_name(db, Nomenclature, values.get("nomenclature", ""))
            if values.get("nomenclature") and product is None:
                errors.append(f"Номенклатура «{values['nomenclature']}» не найдена")
            values["quantity"] = _decimal(values.get("quantity", ""), "Количество", errors)
            values["price"] = _decimal(values.get("price", ""), "Цена", errors)
            action = "stock"

        results.append(RowResult(line=index, values=values, errors=errors, action=action))

    return results


def _ensure_category(db: Session, name: str) -> ProductCategory | None:
    if not name:
        return None
    found = _by_code_or_name(db, ProductCategory, name)
    if found is not None:
        return found
    created = ProductCategory(name=name)
    db.add(created)
    db.flush()
    return created


def _ensure_outlet_type(db: Session, name: str) -> OutletType | None:
    if not name:
        return None
    found = _by_code_or_name(db, OutletType, name)
    if found is not None:
        return found
    created = OutletType(name=name)
    db.add(created)
    db.flush()
    return created


def apply(db: Session, kind: ImportKind, rows: list[RowResult]) -> dict[str, int]:
    """Записывает проверенные строки. Строки с ошибками пропускаются."""

    created = updated = skipped = 0

    for row in rows:
        if not row.valid:
            skipped += 1
            continue

        values = row.values

        if kind is ImportKind.NOMENCLATURE:
            item = _by_code_or_name(db, Nomenclature, values.get("code") or "") or Nomenclature()
            is_new = item.id is None
            item.code = values.get("code") or item.code
            item.name = values["name"]
            item.article = values.get("article") or None
            item.barcode = values.get("barcode") or None
            category = _ensure_category(db, values.get("category", ""))
            if category:
                item.category_id = category.id
            unit = _by_code_or_name(db, UnitOfMeasure, values.get("unit", ""))
            if unit:
                item.base_unit_id = unit.id
            if values.get("price") is not None:
                item.price = values["price"]
            if values.get("vat_rate") is not None:
                item.vat_rate = values["vat_rate"]
            if values.get("is_weight_goods") is not None:
                item.is_weight_goods = values["is_weight_goods"]

        elif kind is ImportKind.COUNTERPARTIES:
            item = _by_code_or_name(db, Counterparty, values.get("code") or "") or Counterparty()
            is_new = item.id is None
            item.code = values.get("code") or item.code
            item.name = values["name"]
            item.bin_iin = values.get("bin_iin") or None
            item.address = values.get("address") or None
            item.phone = values.get("phone") or None
            item.email = values.get("email") or None
            item.contact_person = values.get("contact_person") or None

        elif kind is ImportKind.OUTLETS:
            item = _by_code_or_name(db, Outlet, values.get("code") or "") or Outlet()
            is_new = item.id is None
            counterparty = _by_code_or_name(db, Counterparty, values.get("counterparty", ""))
            item.code = values.get("code") or item.code
            item.name = values["name"]
            if counterparty:
                item.counterparty_id = counterparty.id
            outlet_type = _ensure_outlet_type(db, values.get("outlet_type", ""))
            if outlet_type:
                item.outlet_type_id = outlet_type.id
            item.address = values.get("address") or None
            item.phone = values.get("phone") or None
            item.latitude = values.get("latitude")
            item.longitude = values.get("longitude")

        else:
            # Остатки записываются документом инвентаризации в вызывающем коде.
            continue

        if is_new:
            db.add(item)
            created += 1
        else:
            updated += 1

    db.flush()
    return {"created": created, "updated": updated, "skipped": skipped}
